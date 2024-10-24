import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import os
import time
import re
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image

def next_page():
    global driver
    # 查找具有特定 class 的元素
    elements = driver.find_elements(By.XPATH, "//*[contains(@class, 'sc-d98f2c-0') and contains(@class, 'sc-xhhh7v-2') and contains(@class, 'cCkJiq') and contains(@class, 'sc-xhhh7v-1-filterProps-Styled-Component') and contains(@class, 'kKBslM')]")
    driver.execute_script("arguments[0].scrollIntoView(true);", elements[1])
    elements[1].click()

def excel_title():
    title = ["作品名稱", "作者", "作品"]
    w2xlsx("Pixiv.xlsx", title)

def w2xlsx(file, data):
    try:
        wb = load_workbook(filename=file)
        ws1 = wb["data"]
    except FileNotFoundError:
        wb = Workbook()
        ws1 = wb.create_sheet("data", 0)  # 將 ws1 設定為新建的工作表

    ws1.append(data)
    wb.save(file)

chrome_options = Options()
driver = webdriver.Chrome()

excel_title()


#設置reforer和User-Agent防止被擋
headers = {
    'Referer': 'https://www.pixiv.net/',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
}

url = 'https://www.pixiv.net/tags/%E3%83%8E%E3%83%A9%E3%82%AC%E3%83%9F/illustrations'
driver.get(url)
driver.maximize_window()
count = 0


# 創建目錄存放圖片
if not os.path.exists('images'):
    os.makedirs('images')

for i in range(10):
# 等待頁面載入
    time.sleep(3)

    # 設定滾動等待時間
    scroll_pause_time = 1

    # 獲取當前滾動高度
    last_height = driver.execute_script("return window.scrollY;")
    #設置圖片編號防止相同作品名稱導致下載出錯

    while True:
    # 滾動一小段
        driver.execute_script("window.scrollBy(0, 1000);")
        time.sleep(scroll_pause_time)
        
        # 檢查滾動後的高度
        new_height = driver.execute_script("return window.scrollY;")
        if new_height == last_height:
            break  # 已經到頁面底部

        last_height = new_height
        # 查找所有 img 標籤元素
        #image_elements = driver.find_elements(By.TAG_NAME, 'img')
        image_elements = driver.find_elements(By.XPATH, "//img[contains(@class, 'sc-rp5asc-10') and contains(@class, 'jBxEmj')]")
        
        image_elements = image_elements[4:]

        title = driver.find_elements(By.XPATH,"//*[contains(@class, 'sc-d98f2c-0') and contains(@class, 'sc-iasfms-6') and contains(@class, 'gqlfsh')]")

        author = driver.find_elements(By.XPATH, "//*[contains(@class, 'sc-d98f2c-0') and contains(@class, 'sc-1rx6dmq-2') and contains(@class, 'kghgsn')]")

        # 下載圖片
    for i in range(len(image_elements)):
        data_list = []
        image_url = image_elements[i].get_attribute("src")
        if image_url and ("jpg" in image_url or "png" in image_url):
            img_data = requests.get(image_url, headers=headers).content
            valid_filename = re.sub(r'[<>:"/\\|?*]', '', title[i].text)  # 使用正則表達式替換無效字符
            filename = f'images/{valid_filename}_{count}.jpg'
            
            with open(filename, 'wb') as handler:
                handler.write(img_data)
            data_list.append(valid_filename)
            data_list.append(author[i].text)
            # 插入圖片
            img_path = "images/%s_%d.jpg"%(valid_filename,count)  # 圖片路徑
            wb = load_workbook(filename="Pixiv.xlsx")
            ws1 = wb["data"]

            if os.path.exists(img_path):  # 檢查圖片是否存在
                img = Image(img_path)
                #img.width, img.height = imgsize
                ws1.add_image(img, f'C{count+2}')  # 將圖片插入到指定的單元格
            else:
                print(f"Image not found: {img_path}")  # 如果圖片不存在，則打印提示
            wb.save("Pixiv.xlsx")
            # 將圖片插入到指定儲存格
            w2xlsx("Pixiv.xlsx",data_list)
            count+=1
    try:
        next_page()
    except:
        pass
    #將excel列高調高 已放置圖片
wb = load_workbook(filename="Pixiv.xlsx")
ws1 = wb["data"]
for row in range(1, ws1.max_row + 1):
    ws1.row_dimensions[row].height = 250  # 設置列高
ws1.row_dimensions[1].height = 20
# 關閉瀏覽器
wb.save("Pixiv.xlsx")
driver.quit()
